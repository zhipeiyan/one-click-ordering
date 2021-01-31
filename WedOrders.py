# One click ordering
from itertools import repeat

import openpyxl
import xlrd
from msedge.selenium_tools import Edge, EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait

# configuration of the order file
orders = r'C:\path\to\your.xlsx'
date = 'Jan 1 run4uhome'
columns_per_person = 3
skip_head_lines = 2
skip_rear_lines = 9
website = 'run4uhome'
# website = 'uueat'

# configuration of your system
browser_driver_path = r'C:\lib\edgedriver_win64\msedgedriver.exe'
# browser_driver_path = r'C:\lib\chromedriver_win32\chromedriver.exe'

options = EdgeOptions()
options.use_chromium = True
driver = Edge(options=options, executable_path=browser_driver_path)
# driver = Chrome(executable_path=browser_driver_path)
# driver.maximize_window()
wait = WebDriverWait(driver, 10)


def read_sheet(file, sheet):
    ws = xlrd.open_workbook(file).sheet_by_name(sheet)
    # xlrd can't read hyperlinks from .xlsx files, use openpyxl
    read_links = openpyxl.load_workbook(file)[sheet]
    items = {}
    for person in range(ws.ncols // columns_per_person):
        for row in range(skip_head_lines, ws.nrows - skip_rear_lines):
            # In xlrd, type 2 means number
            if ws.cell(row, columns_per_person * person + 1).ctype == 2:
                # openpyxl indexed from 1, while xlrd indexed from 0
                link = read_links.cell(row=row + 1, column=columns_per_person * person + 1).hyperlink
                count = ws.cell(row, columns_per_person * person + 1).value
                if link is not None:
                    if link.target in items:
                        items[link.target] += count
                    else:
                        items[link.target] = count
                else:
                    print('Error! No url provided for the item:', ws.cell(0, columns_per_person * person).value,
                          ws.cell(row, columns_per_person * person).value)
    return items


def add_to_bag(url, count):
    if count >= 1:
        driver.get(url)
        btn = wait.until(expected_conditions.presence_of_element_located((By.XPATH, "//span[text()='Add to Bag']/..")))
        for _ in repeat(None, count):
            btn.click()


if __name__ == '__main__':
    order = read_sheet(orders, date)
    for item, quantity in order.items():
        if quantity.is_integer():
            add_to_bag(item, int(quantity))
        else:
            print('Error! None integer quantity of an item in the order:', item)
    driver.get('https://www.' + website + '.com/cart')
    input('Press enter to exit.')
    driver.quit()
