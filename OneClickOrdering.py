# One click ordering
import sys
from itertools import repeat

import openpyxl
import xlrd
from msedge.selenium_tools import Edge, EdgeOptions
from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait

# configuration of the order file
order_file = r'C:\path\to\your.xlsx'
# sheet name in your xlsx
date = 'Jan 1 Retailer'
retailer = 'uueat'
# retailer = 'run4uhome'

# configuration of your system
browser = 'Edge'
browser_driver_path = r'C:\lib\edgedriver_win64\msedgedriver.exe'


def init_browser_driver(browser_name, driver_path):
    if browser_name == 'Edge':
        # driver_path = r'C:\lib\edgedriver_win64\msedgedriver.exe'
        options = EdgeOptions()
        options.use_chromium = True
        return Edge(options=options, executable_path=driver_path)
    elif browser_name == 'Chrome':
        # driver_path = r'C:\lib\chromedriver_win32\chromedriver.exe'
        return Chrome(executable_path=driver_path)
    else:
        sys.exit('Implement the browser driver initialization of your choice here.')


def init_website_settings(retailer_name):
    # columns_per_person, skip_head_lines, skip_rear_lines, websites
    # add any retailer powered by Ecwid you want to use to websites
    if retailer_name == 'run4uhome':
        return 3, 2, 7, ['run4uhome']
    elif retailer_name == 'uueat':
        return 4, 2, 14, ['uueat', 'uucart']
    else:
        print('Error! None supported retailer', retailer_name)
        return 0, 0, 0, []


def read_sheet(file, sheet, cols_per_person, skip_head, skip_rear):
    ws = xlrd.open_workbook(file).sheet_by_name(sheet)
    # xlrd can't read hyperlinks from .xlsx files, use openpyxl
    read_links = openpyxl.load_workbook(file)[sheet]
    items = {}
    for person in range(ws.ncols // cols_per_person):
        for row in range(skip_head, ws.nrows - skip_rear):
            # In xlrd, type 2 means number
            if ws.cell(row, cols_per_person * person + cols_per_person - 2).ctype == 2:
                # openpyxl indexed from 1, while xlrd indexed from 0
                link = read_links.cell(row=row + 1, column=cols_per_person * person + 1).hyperlink
                count = ws.cell(row, cols_per_person * person + cols_per_person - 2).value
                if link is not None:
                    if link.target in items:
                        items[link.target] += count
                    else:
                        items[link.target] = count
                else:
                    print('Error! No url provided for the item:', ws.cell(0, cols_per_person * person).value,
                          ws.cell(row, cols_per_person * person).value)
    return items


def add_to_bag(url, count):
    if count >= 1:
        driver.get(url)
        btn = wait.until(expected_conditions.presence_of_element_located((By.XPATH, "//span[text()='Add to Bag']/..")))
        for _ in repeat(None, count):
            btn.click()


if __name__ == '__main__':
    driver = init_browser_driver(browser, browser_driver_path)
    # driver.maximize_window()
    wait = WebDriverWait(driver, 10)

    columns_per_person, skip_head_lines, skip_rear_lines, websites = init_website_settings(retailer)
    orders = read_sheet(order_file, date, columns_per_person, skip_head_lines, skip_rear_lines)

    for item, quantity in orders.items():
        if quantity.is_integer():
            add_to_bag(item, int(quantity))
        else:
            print('Error! None integer quantity of an item in the order:', item)

    if len(websites):
        driver.get('https://www.' + websites[0] + '.com/cart')
    for website in websites[1:]:
        driver.execute_script('window.open("http://www.' + website + '.com/cart","_blank");')

    input('Done. Press enter to exit.')
    driver.quit()
